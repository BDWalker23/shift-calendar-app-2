# Excel calendar generation code placeholder
import calendar
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
import tempfile
import os

def get_color(name):
    if name == "Brandon":
        return "FFFF99"  # Yellow
    elif name == "Tony":
        return "C6EFCE"  # Green
    elif name == "Erik":
        return "BDD7EE"  # Blue
    else:
        return None

def generate_excel_calendar(year, month, schedule):
    wb = Workbook()
    ws = wb.active
    ws.title = f"{calendar.month_name[month]} {year}"

    days = ["Sunday", "Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday"]
    for col, day in enumerate(days, start=1):
        cell = ws.cell(row=1, column=col, value=day)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col)].width = 24

    weeks = calendar.Calendar(firstweekday=6).monthdatescalendar(year, month)
    row_offset = 2
    for row, week in enumerate(weeks, start=row_offset):
        for col, day in enumerate(week, start=1):
            cell = ws.cell(row=row, column=col)
            cell.alignment = Alignment(wrap_text=True, vertical="top")

            if day.month == month:
                off_name = schedule.get(day, "")
                on_names = [n for n in ["Brandon", "Tony", "Erik"] if n != off_name]

                color = get_color(off_name)
                if color:
                    cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

                day_text = f"{day.day}"
                if off_name:
                    cell_value = f"{day_text}\n{off_name} OFF\n{on_names[0]} & {on_names[1]} ON"
                else:
                    cell_value = f"{day_text}"

                # Format day number bold & underline
                parts = cell_value.split("\n")
                cell.value = cell_value
                cell.font = Font(name="Calibri", size=10)
                if parts:
                    cell.font = Font(bold=True, underline="single")
            else:
                cell.value = ""

    ws.page_margins = PageMargins(left=0.25, right=0.25, top=0.5, bottom=0.5)

    # Save as Excel then convert to PDF
    temp_excel = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    wb.save(temp_excel.name)

    # Export to PDF using LibreOffice or similar (example command)
    output_pdf = temp_excel.name.replace(".xlsx", ".pdf")
    os.system(f'libreoffice --headless --convert-to pdf "{temp_excel.name}" --outdir "{os.path.dirname(temp_excel.name)}"')

    return output_pdf
