import streamlit as st
from datetime import datetime, date, timedelta
import calendar
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO

# Define colors for each person's OFF day
COLORS = {
    "Brandon": "FFFF99",  # Light Yellow
    "Tony": "C6EFCE",     # Light Green
    "Erik": "BDD7EE",     # Light Blue
}

def generate_excel_calendar(year, month, schedule):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{calendar.month_name[month]} {year}"

    # Set headers: Sunday - Saturday
    days = ["Sunday", "Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday"]
    for col_num, day_name in enumerate(days, start=1):
        ws.cell(row=1, column=col_num).value = day_name
        ws.cell(row=1, column=col_num).font = Font(bold=True)
        ws.cell(row=1, column=col_num).alignment = Alignment(horizontal="center")

    cal = calendar.Calendar(firstweekday=6)
    weeks = cal.monthdatescalendar(year, month)

    row_offset = 2
    for week_idx, week in enumerate(weeks):
        for day_idx, day in enumerate(week):
            if day.month == month:
                off_person = schedule.get(day)
                on_people = [p for p in ["Brandon", "Tony", "Erik"] if p != off_person]

                cell = ws.cell(row=week_idx + row_offset, column=day_idx + 1)
                cell_value = f"{day.day}\n{off_person} OFF\n{on_people[0]} & {on_people[1]} ON"
                cell.value = cell_value

                # Style
                cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="top")
                if off_person in COLORS:
                    cell.fill = PatternFill(start_color=COLORS[off_person], end_color=COLORS[off_person], fill_type="solid")
                cell.font = Font(size=9)
            else:
                ws.cell(row=week_idx + row_offset, column=day_idx + 1).value = ""

    # Adjust column widths and row heights
    for col in range(1, 8):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20
    for row in range(2, 2 + len(weeks)):
        ws.row_dimensions[row].height = 50

    # Save to in-memory file
    output = BytesIO()
    wb.save(output)
    return output.getvalue()

# Streamlit UI
st.title("Shift Calendar App v2")

selected_year = st.selectbox("Select Year", list(range(2025, 2031)), index=0)
selected_month = st.selectbox("Select Month", list(calendar.month_name)[1:], index=datetime.now().month - 1)

# Manual entry
st.subheader("Select OFF Days")
schedule = {}
cal = calendar.Calendar(firstweekday=6)
weeks = cal.monthdatescalendar(selected_year, selected_month)

for week in weeks:
    for day in week:
        if day.month == selected_month:
            key = f"{day}"
            selected = st.selectbox(f"Off for {day.strftime('%A, %B %d')}", ["", "Brandon", "Tony", "Erik"], key=key)
            if selected:
                schedule[day] = selected

# Auto-fill blanks and balance OFF days
if st.button("Auto-balance schedule"):
    all_days = [d for week in weeks for d in week if d.month == selected_month]
    assigned_days = list(schedule.keys())
    unassigned_days = [d for d in all_days if d not in assigned_days]

    # Count current off days
    from collections import Counter
    count = Counter(schedule.values())

    # Ensure at least one full weekend (Fri–Sun) per person
    weekends = []
    for i in range(len(all_days) - 2):
        if all_days[i].weekday() == 4:  # Friday
            if all(all_days[j].month == selected_month for j in range(i, i + 3)):
                weekends.append(all_days[i:i + 3])

    used_weekends = set()
    for person in ["Brandon", "Tony", "Erik"]:
        for wk in weekends:
            if not any(day in schedule for day in wk):
                for d in wk:
                    schedule[d] = person
                used_weekends.update(wk)
                break

    # Fill the rest evenly
    for d in unassigned_days:
        if d not in used_weekends:
            least = min(count.values()) if count else 0
            candidates = [p for p in ["Brandon", "Tony", "Erik"] if count[p] <= least]
            pick = candidates[0]
            schedule[d] = pick
            count[pick] += 1

# Generate and download
if st.button("Generate Excel Calendar"):
    excel_data = generate_excel_calendar(selected_year, selected_month, schedule)
    st.download_button(
        label="📥 Download Calendar as Excel",
        data=excel_data,
        file_name=f"shift_calendar_{calendar.month_name[selected_month].lower()}_{selected_year}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
