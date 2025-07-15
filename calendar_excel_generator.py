import calendar
from datetime import date, timedelta, datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from collections import Counter
import random

def get_color(name):
    if name == "Brandon":
        return "FFFF99"  # Yellow
    elif name == "Tony":
        return "CCFFCC"  # Green
    elif name == "Erik":
        return "99CCFF"  # Blue
    else:
        return "FFFFFF"  # Default white

def assign_one_full_weekend(person, off_days_set, max_off_days, year, month):
    upgraded = False
    for day in range(1, 29):
        try:
            date_obj = datetime(year, month, day)
        except:
            continue
        if date_obj.weekday() == 4:
            fri = date_obj
            sat = fri + timedelta(days=1)
            sun = fri + timedelta(days=2)
            weekend = {fri, sat, sun}
            selected = weekend & off_days_set
            if selected and not upgraded:
                extra_needed = len(weekend - selected)
                if len(off_days_set) + extra_needed <= max_off_days:
                    off_days_set -= selected
                    off_days_set |= weekend
                    upgraded = True
    return off_days_set

def autopopulate_schedule(year, month, existing_schedule):
    all_names = ["Brandon", "Tony", "Erik"]
    schedule = existing_schedule.copy()
    cal = calendar.Calendar(firstweekday=6)
    all_days = [day for week in cal.monthdatescalendar(year, month) for day in week if day.month == month]

    # Upgrade partial weekends to full weekend OFF
    max_off = len(all_days) // len(all_names) + 1
    for person in all_names:
        person_off_days = {d for d, n in schedule.items() if n == person}
        upgraded_days = assign_one_full_weekend(person, person_off_days, max_off, year, month)
        for d in upgraded_days:
            schedule[d] = person

    # Assign one full weekend to each person if not already done
    weekends = []
    for week in cal.monthdatescalendar(year, month):
        fri, sat, sun = week[calendar.FRIDAY], week[calendar.SATURDAY], week[calendar.SUNDAY]
        if all(d.month == month for d in [fri, sat, sun]):
            weekends.append((fri, sat, sun))

    locked = set(schedule.keys())
    used_weekends = set()
    assigned_weekend_names = set()
    for fri, sat, sun in weekends:
        available_names = [n for n in all_names if n not in assigned_weekend_names]
        if not available_names:
            break
        chosen = available_names[0]
        if all(d not in locked for d in [fri, sat, sun]):
            schedule[fri] = chosen
            schedule[sat] = chosen
            schedule[sun] = chosen
            assigned_weekend_names.add(chosen)
            used_weekends.add((fri, sat, sun))
            locked.update([fri, sat, sun])

    # Fill remaining unassigned days evenly
    off_count = Counter(schedule.values())
    unfilled = [d for d in all_days if d not in schedule]

    base_off = len(all_days) // len(all_names)
    extras = len(all_days) % len(all_names)
    target_off = {n: base_off for n in all_names}
    for i in range(extras):
        target_off[all_names[i]] += 1

    random.shuffle(unfilled)
    for day in unfilled:
        sorted_names = sorted(all_names, key=lambda n: off_count[n])
        for name in sorted_names:
            if off_count[name] < target_off[name]:
                schedule[day] = name
                off_count[name] += 1
                break

    return schedule

def generate_excel_calendar(year, month, schedule, file_path):
    wb = Workbook()
    ws = wb.active
    ws.title = f"{calendar.month_name[month]} {year} Shift Calendar – N969PW"

    title = f"{calendar.month_name[month]} {year} Shift Calendar – N969PW"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = title
    title_cell.font = Font(size=16, bold=True)
    title_cell.alignment = Alignment(horizontal="center")

    days = ["Sunday", "Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday"]
    for col, day in enumerate(days, start=1):
        cell = ws.cell(row=2, column=col)
        cell.value = day
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    cal = calendar.Calendar(firstweekday=6)
    month_days = cal.monthdatescalendar(year, month)

    thin_border = Border(
        top=Side(style="thin"),
        bottom=Side(style="thin"),
        left=Side(style="thin"),
        right=Side(style="thin")
    )

    for week_idx, week in enumerate(month_days):
        for day_idx, day in enumerate(week):
            cell = ws.cell(row=3 + week_idx, column=1 + day_idx)
            if day.month == month:
                off_name = schedule.get(day, "")
                on_names = [n for n in ["Brandon", "Tony", "Erik"] if n != off_name]

                lines = [f"{day.day}"]
                if off_name:
                    lines.append(f"{off_name} OFF")
                    lines.append(f"{on_names[0]} & {on_names[1]} ON")

                cell.value = "\n".join(lines)
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                cell.border = thin_border

                if off_name:
                    fill_color = get_color(off_name)
                    cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

                cell.font = Font(bold=True, underline="single")
            else:
                cell.border = thin_border

    for col_idx in range(1, 8):
        col_letter = chr(64 + col_idx)
        ws.column_dimensions[col_letter].width = 25

    wb.save(file_path)
